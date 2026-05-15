"""
把目前累積的結算報告匯出成 Excel。
資料來源：docs/data/history.json + docs/data/stats.json（由 web_export.py 產生）
輸出：settlement_report_<updated_at>.xlsx
"""
import argparse
import json
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), 'docs', 'data')

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='305496')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=12)
WIN_FILL = PatternFill('solid', fgColor='C6EFCE')
LOSS_FILL = PatternFill('solid', fgColor='FFC7CE')
NEUTRAL_FILL = PatternFill('solid', fgColor='FFEB9C')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _load_json(name):
    with open(os.path.join(DATA_DIR, name), encoding='utf-8') as f:
        return json.load(f)


def _autosize(ws, min_w=8, max_w=24):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = min_w
        for c in col_cells:
            if c.value is None:
                continue
            v = str(c.value)
            w = sum(2 if ord(ch) > 127 else 1 for ch in v)
            length = max(length, w + 2)
        ws.column_dimensions[col_letter].width = min(length, max_w)


def _write_header(ws, row, headers):
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def _fill_for_pct(pct):
    if pct is None:
        return None
    if pct > 0:
        return WIN_FILL
    if pct < 0:
        return LOSS_FILL
    return NEUTRAL_FILL


def _fmt_pct(v):
    return None if v is None else round(float(v), 2)


def sheet_overview(wb, stats, history):
    ws = wb.create_sheet('總覽')
    ws['A1'] = '結算報告 · 總覽'
    ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
    ws.merge_cells('A1:D1')
    ws['A2'] = f"資料時間：{stats.get('updated_at', '-')}"
    ws['A2'].font = Font(italic=True, color='595959')
    ws.merge_cells('A2:D2')

    screen_dates = sorted({r['screen_date'] for r in history.get('records', []) if r.get('screen_date')})
    if screen_dates:
        ws['A3'] = f"涵蓋篩選日：{screen_dates[0]} ~ {screen_dates[-1]}（{len(screen_dates)} 個交易日）"
        ws['A3'].font = Font(italic=True, color='595959')
        ws.merge_cells('A3:D3')

    s = stats.get('summary') or {}
    rows = [
        ('累積篩選數',     s.get('total')),
        ('成功撮合（filled）', s.get('filled')),
        ('未成交（missed）',  s.get('missed')),
        ('尚未撮合（pending）', s.get('pending')),
        ('第一週已結算',  s.get('settled1')),
        ('第一週獲利數',  s.get('win1')),
        ('第一週勝率',    f"{round((s.get('win1') or 0) / s['settled1'] * 100, 1)}%" if s.get('settled1') else '-'),
        ('第一週平均報酬', f"{_fmt_pct(s.get('avg_ret1'))}%" if s.get('avg_ret1') is not None else '-'),
        ('第一週最佳報酬', f"{_fmt_pct(s.get('best_ret1'))}%" if s.get('best_ret1') is not None else '-'),
        ('第一週最差報酬', f"{_fmt_pct(s.get('worst_ret1'))}%" if s.get('worst_ret1') is not None else '-'),
        ('第二週已結算',   s.get('settled2')),
        ('第二週獲利數',   s.get('win2')),
        ('第二週勝率',     f"{round((s.get('win2') or 0) / s['settled2'] * 100, 1)}%" if s.get('settled2') else '-'),
        ('第二週平均報酬', f"{_fmt_pct(s.get('avg_ret2'))}%" if s.get('avg_ret2') is not None else '-'),
    ]
    start = 5
    ws.cell(row=start, column=1, value='指標').fill = HEADER_FILL
    ws.cell(row=start, column=1).font = HEADER_FONT
    ws.cell(row=start, column=2, value='數值').fill = HEADER_FILL
    ws.cell(row=start, column=2).font = HEADER_FONT
    for i, (k, v) in enumerate(rows, start=start + 1):
        ws.cell(row=i, column=1, value=k).border = BORDER
        c = ws.cell(row=i, column=2, value=v)
        c.border = BORDER
        c.alignment = Alignment(horizontal='right')

    mh = stats.get('missed_hypo') or {}
    if mh.get('total'):
        r = start + len(rows) + 3
        ws.cell(row=r, column=1, value='Missed 假設結算（若有買到的話）').font = Font(bold=True, color='1F4E78')
        r += 1
        for k, label in [('total', '樣本數'), ('win', '獲利數'),
                         ('win_rate', '勝率(%)'), ('avg_ret', '平均報酬(%)'),
                         ('best', '最佳報酬(%)'), ('worst', '最差報酬(%)')]:
            ws.cell(row=r, column=1, value=label).border = BORDER
            ws.cell(row=r, column=2, value=mh.get(k)).border = BORDER
            r += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18


def sheet_records(wb, history):
    ws = wb.create_sheet('篩選明細')
    headers = [
        '篩選日', '股號', '名稱', '等級', '評分',
        '收盤價', '漲幅%', '量比', '乖離%', '追價模式', '連續漲停',
        '進場區間下緣', '進場區間上緣', '撮合狀態',
        '實際進場日', '實際進場價', '目標1', '目標2', '停損',
        'W1 報酬%', 'W1 結算', 'W2 報酬%', 'W2 結算',
        '觸目標1', '觸目標2', '觸停損',
        '觸目標1日', '觸目標2日', '觸停損日',
    ]
    _write_header(ws, 1, headers)
    ws.freeze_panes = 'D2'

    keys = [
        'screen_date', 'sid', 'name', 'grade', 'score',
        'close_price', 'change_pct', 'vol_ratio', 'bias_pct', 'chase_mode', 'consec_limit_up',
        'entry_zone_low', 'entry_zone_high', 'fill_status',
        'actual_entry_date', 'actual_entry_price', 'actual_target1', 'actual_target2', 'actual_stop_loss',
        'settle1_pct', 'settle1_done', 'settle2_pct', 'settle2_done',
        'hit_target1', 'hit_target2', 'hit_stoploss',
        'hit_target1_date', 'hit_target2_date', 'hit_stoploss_date',
    ]
    records = history.get('records', [])
    records_sorted = sorted(
        records, key=lambda r: (r.get('screen_date') or '', r.get('grade') or '', r.get('sid') or '')
    )
    for i, r in enumerate(records_sorted, start=2):
        for col, key in enumerate(keys, 1):
            v = r.get(key)
            if isinstance(v, bool):
                v = '是' if v else '否'
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        # 套色：以 settle1_pct 為主
        fill = _fill_for_pct(r.get('settle1_pct'))
        if fill is not None:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill

    _autosize(ws)


def sheet_by_grade(wb, stats):
    ws = wb.create_sheet('各等級績效')
    headers = ['等級', '篩選總數', '成交', '未成交', '尚待撮合',
               'W1 已結算', 'W1 獲利數', 'W1 勝率%', 'W1 平均報酬%',
               'W2 已結算', 'W2 獲利數', 'W2 勝率%', 'W2 平均報酬%',
               '觸目標1', '觸目標2', '觸停損']
    _write_header(ws, 1, headers)
    for i, g in enumerate(stats.get('by_grade', []), start=2):
        s1 = g.get('settled1') or 0
        s2 = g.get('settled2') or 0
        wr1 = round((g.get('win1') or 0) / s1 * 100, 1) if s1 else None
        wr2 = round((g.get('win2') or 0) / s2 * 100, 1) if s2 else None
        row = [
            g.get('grade'), g.get('total'), g.get('filled'), g.get('missed'), g.get('pending'),
            s1, g.get('win1'), wr1, _fmt_pct(g.get('avg_ret1')),
            s2, g.get('win2'), wr2, _fmt_pct(g.get('avg_ret2')),
            g.get('hit_t1'), g.get('hit_t2'), g.get('hit_sl'),
        ]
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center')
        fill = _fill_for_pct(g.get('avg_ret1'))
        if fill is not None:
            ws.cell(row=i, column=9).fill = fill
    _autosize(ws)


def sheet_by_bias(wb, stats):
    ws = wb.create_sheet('乖離率區間')
    headers = ['乖離率區間', '篩選總數', '已結算', '獲利數', '勝率%', '平均報酬%']
    _write_header(ws, 1, headers)
    for i, b in enumerate(stats.get('by_bias', []), start=2):
        settled = b.get('settled') or 0
        wr = round((b.get('win') or 0) / settled * 100, 1) if settled else None
        row = [b.get('bias_zone'), b.get('total'), settled, b.get('win'), wr, _fmt_pct(b.get('avg_ret'))]
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center')
        fill = _fill_for_pct(b.get('avg_ret'))
        if fill is not None:
            ws.cell(row=i, column=6).fill = fill
    _autosize(ws)


def sheet_by_month(wb, stats):
    ws = wb.create_sheet('月度績效')
    headers = ['月份', '篩選總數', '已結算', '獲利數', '勝率%', '平均報酬%']
    _write_header(ws, 1, headers)
    for i, m in enumerate(stats.get('by_month', []), start=2):
        settled = m.get('settled') or 0
        wr = round((m.get('win') or 0) / settled * 100, 1) if settled else None
        row = [m.get('ym'), m.get('total'), settled, m.get('win'), wr, _fmt_pct(m.get('avg_ret'))]
        for col, v in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center')
        fill = _fill_for_pct(m.get('avg_ret'))
        if fill is not None:
            ws.cell(row=i, column=6).fill = fill
    _autosize(ws)


def sheet_timeline(wb, stats):
    ws = wb.create_sheet('結算時序')
    tl = stats.get('timeline') or {}
    headers = ['輪次', '結算日', '結算筆數', '獲利數', '勝率%', '平均報酬%']
    _write_header(ws, 1, headers)
    row_i = 2
    for series_key, label in [('w1', '第一週(W1)'), ('w2', '第二週(W2)')]:
        for r in tl.get(series_key, []):
            total = r.get('total') or 0
            wr = round((r.get('wins') or 0) / total * 100, 1) if total else None
            row = [label, r.get('sdate'), total, r.get('wins'), wr, _fmt_pct(r.get('avg_ret'))]
            for col, v in enumerate(row, 1):
                c = ws.cell(row=row_i, column=col, value=v)
                c.border = BORDER
                c.alignment = Alignment(horizontal='center')
            fill = _fill_for_pct(r.get('avg_ret'))
            if fill is not None:
                ws.cell(row=row_i, column=6).fill = fill
            row_i += 1
    _autosize(ws)


def sheet_missed_hypo(wb, stats):
    mh = stats.get('missed_hypo') or {}
    if not mh.get('total'):
        return
    ws = wb.create_sheet('Missed假設結算')
    ws['A1'] = 'Missed 假設結算（若有買到的話）'
    ws['A1'].font = Font(bold=True, size=13, color='1F4E78')
    ws.merge_cells('A1:F1')

    headers = ['指標', '數值']
    _write_header(ws, 3, headers)
    rows = [
        ('樣本數', mh.get('total')),
        ('獲利數', mh.get('win')),
        ('勝率(%)', mh.get('win_rate')),
        ('平均報酬(%)', mh.get('avg_ret')),
        ('最佳報酬(%)', mh.get('best')),
        ('最差報酬(%)', mh.get('worst')),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).border = BORDER
        ws.cell(row=i, column=2, value=v).border = BORDER

    by_grade = mh.get('by_grade') or []
    if by_grade:
        r0 = 4 + len(rows) + 2
        ws.cell(row=r0, column=1, value='各等級').font = Font(bold=True, color='1F4E78')
        _write_header(ws, r0 + 1, ['等級', '樣本數', '獲利數', '勝率%', '平均報酬%'])
        for i, g in enumerate(by_grade, start=r0 + 2):
            row = [g.get('grade'), g.get('total'), g.get('win'),
                   g.get('win_rate'), g.get('avg_ret')]
            for col, v in enumerate(row, 1):
                c = ws.cell(row=i, column=col, value=v)
                c.border = BORDER
                c.alignment = Alignment(horizontal='center')
    _autosize(ws)


def build_workbook(out_path):
    stats = _load_json('stats.json')
    history = _load_json('history.json')

    wb = Workbook()
    wb.remove(wb.active)

    sheet_overview(wb, stats, history)
    sheet_records(wb, history)
    sheet_by_grade(wb, stats)
    sheet_by_bias(wb, stats)
    sheet_by_month(wb, stats)
    sheet_timeline(wb, stats)
    sheet_missed_hypo(wb, stats)

    wb.save(out_path)
    return out_path


def _default_filename(stats_updated_at):
    ts = re.sub(r'[^0-9]', '', stats_updated_at or '')[:12]
    if not ts:
        ts = datetime.now().strftime('%Y%m%d%H%M')
    return f'settlement_report_{ts}.xlsx'


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-o', '--output', help='輸出檔名（預設 settlement_report_<時間>.xlsx）')
    args = parser.parse_args()

    stats = _load_json('stats.json')
    out = args.output or _default_filename(stats.get('updated_at'))
    path = os.path.abspath(out)
    build_workbook(path)
    print(f'已輸出：{path}')


if __name__ == '__main__':
    main()
